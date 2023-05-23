from flask import Flask, render_template, request, redirect, url_for, send_file
import os
import tempfile
import ifcopenshell
import ifcopenshell.util.element
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

def load_element_types(file):
    element_types = {}
    df = pd.read_excel(file, sheet_name="Element_Types", engine='openpyxl')
    for index, row in df.iterrows():
        ifc_class = row["IFC_Class"]
        element_types[ifc_class] = ifc_class
    return element_types

def load_required_psets_and_params(file):
    required_psets_and_params = {}
    xls = pd.ExcelFile(file, engine='openpyxl')
    for sheet_name in xls.sheet_names:
        if sheet_name != "Element_Types":
            df = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
            for index, row in df.iterrows():
                ifc_class = row["IFC_Class"]
                param_name = row["Parametre"]
                param_type = row["Type"]
                if ifc_class not in required_psets_and_params:
                    required_psets_and_params[ifc_class] = {}
                if sheet_name not in required_psets_and_params[ifc_class]:
                    required_psets_and_params[ifc_class][sheet_name] = {}
                required_psets_and_params[ifc_class][sheet_name][param_name] = param_type
    return required_psets_and_params

def str_to_type(param_type_str):
    param_type_str = param_type_str.lower()
    if param_type_str.lower() == 'string':
        return str
    elif param_type_str.lower() == 'int':
        return int
    elif param_type_str.lower() == 'float':
        return float
    elif param_type_str == 'number':
        return float
    elif param_type_str.lower() == 'bool':
        return bool
    else:
        raise ValueError(f"Unsupported parameter type '{param_type_str}'")

def gray_empty_cells(worksheet):
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None or cell.value == "":
                cell.fill = gray_fill

def process_files(temp_dir, ifc_file_path, excel_file_path, output_file_path):
    element_types = load_element_types(excel_file_path)
    required_psets_and_params = load_required_psets_and_params(excel_file_path)

    # Load the IFC file
    ifc_file = ifcopenshell.open(ifc_file_path)

    # Get the list of elements and filter by element type
    elements = ifc_file.by_type("IfcElement")
    filtered_elements = [e for e in elements if e.is_a() in element_types]

    # Creating an output Excel workbook
    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)  # Remove the default active sheet

    pset_sheets = {}  # A dictionary to store the sheets created for each PSet

    for ifc_class, psets in required_psets_and_params.items():
        for pset_name in psets:
            # Create a new sheet for each PSet if it doesn't exist yet
            if pset_name not in pset_sheets:
                pset_sheets[pset_name] = output_workbook.create_sheet(f"Results_{pset_name}")
                pset_sheets[pset_name].append(["IFC Class", "Element GlobalId", "Pset Name", "Parameter Name", "Value", "Status"])

            for element in filtered_elements:
                if element.is_a() == ifc_class:
                    pset = ifcopenshell.util.element.get_psets(element).get(pset_name, None)

                    for param_name, param_type_str in required_psets_and_params[ifc_class][pset_name].items():
                        param_type = str_to_type(param_type_str)
                        if pset is None:
                            pset_sheets[pset_name].append([ifc_class, element.GlobalId, pset_name, param_name, "", "PSet missing"])
                        else:
                            value = pset.get(param_name, None)
                            if value is None:
                                pset_sheets[pset_name].append([ifc_class, element.GlobalId, pset_name, param_name, "", "Param missing"])
                            else:
                                pset_sheets[pset_name].append([ifc_class, element.GlobalId, pset_name, param_name, value, "OK"])

    # Create an "Overview" sheet
    overview_sheet = output_workbook.create_sheet("Overview")
    overview_sheet.append(["Pset Name", "OK Params", "Missing Params", "Missing PSets"])

    # Count OK params, missing params, and missing PSets
    for sheet in pset_sheets.values():
        ok_params = 0
        missing_params = 0
        missing_ps = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header row
            if row[5] == "OK":
                ok_params += 1
            elif row[5] == "Param missing":
                missing_params += 1
            elif row[5] == "PSet missing":
                missing_ps += 1
        overview_sheet.append([sheet.title, ok_params, missing_params, missing_ps])

    # Apply the formatting
    for sheet in output_workbook:
        gray_empty_cells(sheet)

    # Save the output workbook
    output_workbook.save(output_file_path)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if request.method == 'POST':
        ifc_file = request.files['ifc_file']
        excel_file = request.files['excel_file']
        if ifc_file and excel_file:
            ifc_filename = secure_filename(ifc_file.filename)
            excel_filename = secure_filename(excel_file.filename)
            ifc_file_path = os.path.join(app.config['UPLOAD_FOLDER'], ifc_filename)
            excel_file_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
            ifc_file.save(ifc_file_path)
            excel_file.save(excel_file_path)

            # Extract the base name of the IFC file without the extension
            ifc_base_name = os.path.splitext(ifc_filename)[0]

            with tempfile.TemporaryDirectory() as temp_dir:
                output_file_path = os.path.join(temp_dir, f'output_{ifc_base_name}.xlsx')
                process_files(temp_dir, ifc_file_path, excel_file_path, output_file_path)
                return send_file(output_file_path, as_attachment=True, download_name=f'output_{ifc_base_name}.xlsx')

    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
