from flask import Flask, render_template, request, redirect, url_for, send_file, abort, jsonify
import os
import tempfile
import ifcopenshell
import ifcopenshell.util.element
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import uuid
from collections import defaultdict
import concurrent.futures
from typing import Dict, List, Set, Tuple
import time
import shutil
import re
from datetime import datetime
import threading

app = Flask(__name__)

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 300 * 1024 * 1024  # 300 Mo
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
TEMP_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')

# Créer les dossiers nécessaires
print(f"Creating folders: {UPLOAD_FOLDER}, {TEMP_FOLDER}")
for folder in [UPLOAD_FOLDER, TEMP_FOLDER]:
    if not os.path.exists(folder):
        try:
            os.makedirs(folder, mode=0o777, exist_ok=True)
            print(f"Created folder: {folder}")
        except Exception as e:
            print(f"Error creating folder {folder}: {str(e)}")

ALLOWED_EXTENSIONS = {'ifc', 'xlsx'}

def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def load_element_types(file):
    element_types = {}
    df = pd.read_excel(file, sheet_name="Element_Types", engine='openpyxl')
    for index, row in df.iterrows():
        ifc_class = row["IFC_Class"]
        element_types[ifc_class] = ifc_class
    return element_types

def load_required_psets_and_params(file):
    required_psets_and_params = {}
    xls = pd.ExcelFile(file, engine='openpyxl')
    for sheet_name in xls.sheet_names:
        if sheet_name != "Element_Types":
            df = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
            for index, row in df.iterrows():
                ifc_class = row["IFC_Class"]
                param_name = row["Parametre"]
                param_type = row["Type"]
                if ifc_class not in required_psets_and_params:
                    required_psets_and_params[ifc_class] = {}
                if sheet_name not in required_psets_and_params[ifc_class]:
                    required_psets_and_params[ifc_class][sheet_name] = {}
                required_psets_and_params[ifc_class][sheet_name][param_name] = param_type
    return required_psets_and_params

def str_to_type(param_type_str):
    param_type_str = param_type_str.lower()
    if param_type_str.lower() == 'string':
        return str
    elif param_type_str.lower() == 'int':
        return int
    elif param_type_str.lower() == 'float':
        return float
    elif param_type_str == 'number':
        return float
    elif param_type_str.lower() == 'bool':
        return bool
    else:
        raise ValueError(f"Unsupported parameter type '{param_type_str}'")

def gray_empty_cells(worksheet):
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None or cell.value == "":
                cell.fill = gray_fill

def get_building_storey(ifc_file, element):
    try:
        if element is not None:
            if hasattr(element, 'ContainedInStructure') and element.ContainedInStructure:
                for rel in element.ContainedInStructure:
                    if rel.RelatingStructure.is_a("IfcBuildingStorey"):
                        return rel.RelatingStructure.Name if rel.RelatingStructure.Name else "Sans étage"
                    elif rel.RelatingStructure.is_a("IfcBuilding") or rel.RelatingStructure.is_a("IfcSite"):
                        return "Sans étage"
                    else:
                        return get_building_storey(ifc_file, rel.RelatingStructure)
        return "Sans étage"
    except Exception as e:
        print(f"Error getting building storey: {str(e)}")
        return "Sans étage"

def sort_floor_name(floor_name):
    if floor_name == "Sans étage":
        return "ZZZ"
    elif floor_name == "Sous-sol":
        return "AAA_-1"
    elif floor_name.startswith("R+"):
        try:
            return f"CCC_{int(floor_name[2:]):02d}"
        except ValueError:
            return f"DDD_{floor_name}"
    elif floor_name == "Rez-de-chaussée":
        return "BBB_00"
    elif floor_name.startswith("-"):
        try:
            return f"AAA_{int(floor_name.split('.')[0]):02d}"
        except ValueError:
            return f"DDD_{floor_name}"
    else:
        return f"DDD_{floor_name}"

def create_summary_sheet(workbook, total_elements, valid_elements, missing_elements, missing_psets, missing_params, floor_stats, elements_by_class, required_psets_and_params):
    # Créer l'onglet de résumé
    summary = workbook.create_sheet("Résumé", 0)
    
    # Styles
    title_font = Font(size=14, bold=True, color="FFFFFF")
    header_font = Font(size=12, bold=True)
    normal_font = Font(size=11)
    
    title_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ko_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    percent_format = '0.0%'
    
    def apply_style_to_range(start_cell, end_cell, font=None, fill=None, alignment=None, border=None, value=None, number_format=None):
        start_col = ord(start_cell[0].upper()) - ord('A') + 1
        end_col = ord(end_cell[0].upper()) - ord('A') + 1
        start_row = int(start_cell[1:])
        end_row = int(end_cell[1:])
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = summary.cell(row=row, column=col)
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
                if border:
                    cell.border = border
                if number_format:
                    cell.number_format = number_format
                if value and col == start_col and row == start_row:
                    cell.value = value
    
    # Titre principal
    apply_style_to_range('A1', 'E1', font=title_font, fill=title_fill, alignment=center_alignment, border=border, value="Rapport de validation IFC")
    summary.merge_cells('A1:E1')
    
    # Statistiques globales
    apply_style_to_range('A3', 'A3', font=header_font, fill=header_fill, value="Statistiques globales")
    
    stats = [
        ("Total d'éléments", total_elements),
        ("Éléments valides", valid_elements, valid_elements/total_elements if total_elements > 0 else 0),
        ("Éléments invalides", missing_elements, missing_elements/total_elements if total_elements > 0 else 0),
        ("PSet manquants", missing_psets),
        ("Paramètres manquants", missing_params)
    ]
    
    for i, stat in enumerate(stats, 4):
        label = stat[0]
        value = stat[1]
        percent = stat[2] if len(stat) > 2 else None
        
        apply_style_to_range(f'A{i}', f'A{i}', font=normal_font, alignment=left_alignment, border=border, value=label)
        apply_style_to_range(f'B{i}', f'B{i}', font=normal_font, alignment=center_alignment, border=border, value=value)
        if percent is not None:
            apply_style_to_range(f'C{i}', f'C{i}', font=normal_font, alignment=center_alignment, border=border, value=percent, number_format=percent_format)
    
    # Statistiques par étage
    apply_style_to_range('A7', 'A7', font=header_font, fill=header_fill, value="Répartition par étage")
    
    headers = ["Étage", "Total", "Valides", "Invalides", "Taux de validité"]
    for col, header in enumerate(headers, 1):
        apply_style_to_range(f'{chr(ord("A")+col-1)}8', f'{chr(ord("A")+col-1)}8', 
                           font=header_font, fill=header_fill, alignment=center_alignment, border=border, value=header)
    
    row = 9
    sorted_floors = sorted(floor_stats.items(), key=lambda x: sort_floor_name(x[0]))
    for floor, stats in sorted_floors:
        total = stats["valid"] + stats["invalid"]
        valid_rate = stats["valid"] / total if total > 0 else 0
        
        cells = [
            (floor, left_alignment),
            (total, center_alignment),
            (stats["valid"], center_alignment),
            (stats["invalid"], center_alignment),
            (valid_rate, center_alignment)
        ]
        
        for col, (value, alignment) in enumerate(cells, 1):
            cell_ref = f'{chr(ord("A")+col-1)}{row}'
            fill = None
            number_format = None
            
            if col == 5:  # Dernière colonne (taux)
                number_format = percent_format
                if valid_rate >= 0.8:
                    fill = ok_fill
                elif valid_rate < 0.5:
                    fill = ko_fill
            
            apply_style_to_range(cell_ref, cell_ref, 
                               font=normal_font, 
                               alignment=alignment, 
                               border=border, 
                               fill=fill, 
                               value=value,
                               number_format=number_format)
        
        row += 1
    
    # Statistiques par type
    type_title_row = row + 2
    apply_style_to_range(f'A{type_title_row}', f'A{type_title_row}', font=header_font, fill=header_fill, value="Répartition par type")
    
    headers = ["Type", "Total", "Valides", "Invalides", "Taux de validité"]
    header_row = type_title_row + 1
    for col, header in enumerate(headers, 1):
        apply_style_to_range(f'{chr(ord("A")+col-1)}{header_row}', f'{chr(ord("A")+col-1)}{header_row}',
                           font=header_font, fill=header_fill, alignment=center_alignment, border=border, value=header)
    
    row = header_row + 1
    sorted_types = sorted(elements_by_class.items())
    for element_type, stats in sorted_types:
        total = stats["total"]
        valid_rate = stats["valid"] / total if total > 0 else 0
        
        cells = [
            (element_type, left_alignment),
            (total, center_alignment),
            (stats["valid"], center_alignment),
            (stats["invalid"], center_alignment),
            (valid_rate, center_alignment)
        ]
        
        for col, (value, alignment) in enumerate(cells, 1):
            cell_ref = f'{chr(ord("A")+col-1)}{row}'
            fill = None
            number_format = None
            
            if col == 5:  # Dernière colonne (taux)
                number_format = percent_format
                if valid_rate >= 0.8:
                    fill = ok_fill
                elif valid_rate < 0.5:
                    fill = ko_fill
            
            apply_style_to_range(cell_ref, cell_ref,
                               font=normal_font,
                               alignment=alignment,
                               border=border,
                               fill=fill,
                               value=value,
                               number_format=number_format)
        
        row += 1
    
    # Ajuster la largeur des colonnes
    for col in range(1, 6):
        letter = get_column_letter(col)
        max_length = 0
        for cell in summary[letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        summary.column_dimensions[letter].width = adjusted_width
    
    # Figer les volets
    summary.freeze_panes = 'A2'
    
    # Activer l'onglet de résumé
    workbook.active = workbook.worksheets.index(summary)

def process_files(temp_dir: str, ifc_file_path: str, excel_file_path: str, output_file_path: str):
    start_time = time.time()
    print(f"Starting analysis...")
    
    # Chargement des données
    element_types = load_element_types(excel_file_path)
    required_psets_and_params = load_required_psets_and_params(excel_file_path)
    ifc_file = ifcopenshell.open(ifc_file_path)
    model_name = os.path.basename(ifc_file_path)
    
    # Préfiltrage des éléments
    print(f"Loading elements... ({time.time() - start_time:.2f}s)")
    elements_by_class = {}
    missing_psets = 0
    missing_params = 0
    
    # Statistiques par étage
    floor_stats = defaultdict(lambda: {"valid": 0, "invalid": 0})
    
    # Analyse des éléments
    print(f"Analyzing elements... ({time.time() - start_time:.2f}s)")
    
    # Préparation du rapport Excel
    workbook = Workbook()
    details = workbook.active
    details.title = "Détails"
    
    # Styles pour l'onglet de détails
    header_font = Font(size=12, bold=True)
    normal_font = Font(size=11)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ko_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    no_fill = PatternFill(fill_type=None)
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes de l'onglet de détails
    headers = ["Type", "Étage", "ID", "Nom", "PSet", "Paramètre", "Valeur", "Statut"]
    for col, header in enumerate(headers, 1):
        cell = details.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
        cell.font = header_font
    
    row = 2
    for element in ifc_file.by_type("IfcElement"):
        if element.is_a() in element_types:
            if element.is_a() not in elements_by_class:
                elements_by_class[element.is_a()] = {"total": 0, "valid": 0, "invalid": 0}
            
            elements_by_class[element.is_a()]["total"] += 1
            floor = get_building_storey(ifc_file, element)
            element_psets = ifcopenshell.util.element.get_psets(element)
            element_valid = True
            has_missing_pset = False
            has_missing_param = False
            
            # Vérification des PSet et paramètres requis
            for pset_name, params in required_psets_and_params[element.is_a()].items():
                if pset_name not in element_psets:
                    element_valid = False
                    has_missing_pset = True
                    
                    cells = [
                        (1, element.is_a(), left_alignment),
                        (2, str(floor), left_alignment),
                        (3, element.GlobalId, center_alignment),
                        (4, getattr(element, 'Name', ''), left_alignment),
                        (5, pset_name, left_alignment),
                        (6, "TOUS", center_alignment),
                        (7, "MANQUANT", center_alignment),
                        (8, "KO", center_alignment)
                    ]
                    
                    for col, value, alignment in cells:
                        cell = details.cell(row=row, column=col)
                        cell.value = value
                        cell.alignment = alignment
                        cell.border = border
                        cell.font = normal_font
                        cell.fill = ko_fill if col == 8 else no_fill
                    
                    row += 1
                else:
                    for param_name, param_type in params.items():
                        actual_value = element_psets[pset_name].get(param_name)
                        
                        if actual_value is None:
                            element_valid = False
                            has_missing_param = True
                            
                            cells = [
                                (1, element.is_a(), left_alignment),
                                (2, str(floor), left_alignment),
                                (3, element.GlobalId, center_alignment),
                                (4, getattr(element, 'Name', ''), left_alignment),
                                (5, pset_name, left_alignment),
                                (6, param_name, left_alignment),
                                (7, "MANQUANT", center_alignment),
                                (8, "KO", center_alignment)
                            ]
                            
                            for col, value, alignment in cells:
                                cell = details.cell(row=row, column=col)
                                cell.value = value
                                cell.alignment = alignment
                                cell.border = border
                                cell.font = normal_font
                                cell.fill = ko_fill if col == 8 else no_fill
                            
                            row += 1
                        else:
                            try:
                                param_type_class = str_to_type(param_type)
                                param_type_class(actual_value)
                                
                                cells = [
                                    (1, element.is_a(), left_alignment),
                                    (2, str(floor), left_alignment),
                                    (3, element.GlobalId, center_alignment),
                                    (4, getattr(element, 'Name', ''), left_alignment),
                                    (5, pset_name, left_alignment),
                                    (6, param_name, left_alignment),
                                    (7, str(actual_value), left_alignment),
                                    (8, "OK", center_alignment)
                                ]
                                
                                for col, value, alignment in cells:
                                    cell = details.cell(row=row, column=col)
                                    cell.value = value
                                    cell.alignment = alignment
                                    cell.border = border
                                    cell.font = normal_font
                                    cell.fill = ok_fill if col == 8 else no_fill
                                
                                row += 1
                            except (ValueError, TypeError):
                                element_valid = False
                                has_missing_param = True
                                
                                cells = [
                                    (1, element.is_a(), left_alignment),
                                    (2, str(floor), left_alignment),
                                    (3, element.GlobalId, center_alignment),
                                    (4, getattr(element, 'Name', ''), left_alignment),
                                    (5, pset_name, left_alignment),
                                    (6, param_name, left_alignment),
                                    (7, f"{actual_value} (attendu: {param_type})", left_alignment),
                                    (8, "KO", center_alignment)
                                ]
                                
                                for col, value, alignment in cells:
                                    cell = details.cell(row=row, column=col)
                                    cell.value = value
                                    cell.alignment = alignment
                                    cell.border = border
                                    cell.font = normal_font
                                    cell.fill = ko_fill if col == 8 else no_fill
                                
                                row += 1
            
            # Mise à jour des statistiques
            if element_valid:
                elements_by_class[element.is_a()]["valid"] += 1
                floor_stats[floor]["valid"] += 1
            else:
                elements_by_class[element.is_a()]["invalid"] += 1
                floor_stats[floor]["invalid"] += 1
                if has_missing_pset:
                    missing_psets += 1
                if has_missing_param:
                    missing_params += 1
    
    # Statistiques globales
    total_elements = sum(stats["total"] for stats in elements_by_class.values())
    valid_elements = sum(stats["valid"] for stats in elements_by_class.values())
    invalid_elements = sum(stats["invalid"] for stats in elements_by_class.values())
    
    # Créer l'onglet de résumé
    create_summary_sheet(workbook, total_elements, valid_elements, invalid_elements, missing_psets, missing_params, floor_stats, elements_by_class, required_psets_and_params)
    
    # Ajuster les largeurs des colonnes
    for ws in workbook.worksheets:
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # Figer les volets dans l'onglet de détails
    details.freeze_panes = 'A2'
    
    # Sauvegarder le fichier
    print(f"Saving file... ({time.time() - start_time:.2f}s)")
    workbook.save(output_file_path)
    print(f"Analysis completed in {time.time() - start_time:.2f}s")
    
    # Préparation des données pour le dashboard
    sorted_floors = sorted(floor_stats.items(), key=lambda x: sort_floor_name(x[0]))
    return {
        "total_elements": total_elements,
        "valid_elements": valid_elements,
        "missing_elements": invalid_elements,
        "missing_psets": missing_psets,
        "missing_params": missing_params,
        "floors": [{"name": floor, "valid": stats["valid"], "invalid": stats["invalid"]} for floor, stats in sorted_floors]
    }

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    try:
        print("Starting upload...")
        if 'ifc_file' not in request.files or 'excel_file' not in request.files:
            print("Missing files in request")
            return jsonify({"error": "Missing file"}), 400
        
        ifc_file = request.files['ifc_file']
        excel_file = request.files['excel_file']
        
        print(f"Received files: IFC={ifc_file.filename}, Excel={excel_file.filename}")
        
        if ifc_file.filename == '' or excel_file.filename == '':
            print("Empty filenames")
            return jsonify({"error": "No selected file"}), 400
        
        if not allowed_file(ifc_file.filename) or not allowed_file(excel_file.filename):
            print("Invalid file types")
            return jsonify({"error": "Invalid file type"}), 400
        
        # Générer un ID unique pour cette analyse
        analysis_id = str(uuid.uuid4())
        analysis_dir = os.path.join(TEMP_FOLDER, analysis_id)
        print(f"Creating analysis directory: {analysis_dir}")
        try:
            os.makedirs(analysis_dir, exist_ok=True)
        except Exception as e:
            print(f"Error creating analysis directory: {str(e)}")
            return jsonify({"error": f"Could not create analysis directory: {str(e)}"}), 500
        
        # Sauvegarder les fichiers
        ifc_path = os.path.join(analysis_dir, secure_filename(ifc_file.filename))
        excel_path = os.path.join(analysis_dir, secure_filename(excel_file.filename))
        output_path = os.path.join(analysis_dir, f'output_{os.path.splitext(ifc_file.filename)[0]}.xlsx')
        
        print(f"Saving files to: {ifc_path}, {excel_path}")
        try:
            ifc_file.save(ifc_path)
            excel_file.save(excel_path)
        except Exception as e:
            print(f"Error saving files: {str(e)}")
            return jsonify({"error": f"Could not save files: {str(e)}"}), 500
        
        # Analyser les fichiers
        print("Starting analysis...")
        try:
            results = process_files(analysis_dir, ifc_path, excel_path, output_path)
            results["analysis_id"] = analysis_id
            print("Analysis completed successfully")
            return jsonify(results)
        except Exception as e:
            print(f"Error during analysis: {str(e)}")
            return jsonify({"error": f"Analysis failed: {str(e)}"}), 500
        
    except Exception as e:
        print(f"Unexpected error during upload: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/download/<analysis_id>')
def download(analysis_id):
    try:
        analysis_dir = os.path.join(TEMP_FOLDER, analysis_id)
        if not os.path.exists(analysis_dir):
            return jsonify({"error": "Analysis not found"}), 404
        
        # Trouver le fichier Excel de sortie
        output_files = [f for f in os.listdir(analysis_dir) if f.startswith('output_') and f.endswith('.xlsx')]
        if not output_files:
            return jsonify({"error": "Output file not found"}), 404
        
        output_path = os.path.join(analysis_dir, output_files[0])
        return send_file(output_path, as_attachment=True, download_name=output_files[0])
        
    except Exception as e:
        print(f"Error during download: {str(e)}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    # Configuration du port via variable d'environnement pour Vercel
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)
