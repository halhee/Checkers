from flask import Flask, render_template, request, redirect, url_for, send_file, abort, jsonify
import os
import tempfile
import ifcopenshell
import ifcopenshell.util.element
import uuid
import time
import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import concurrent.futures
from typing import Dict, List, Set, Tuple
import shutil
import re
from datetime import datetime
import threading
from werkzeug.utils import secure_filename
import folium

app = Flask(__name__)

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 300 * 1024 * 1024  # 300 Mo
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
TEMP_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
GEOMAP_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'geomaps')
os.makedirs(GEOMAP_FOLDER, exist_ok=True)


# Créer les dossiers nécessaires
print(f"Creating folders: {UPLOAD_FOLDER}, {TEMP_FOLDER}")
for folder in [UPLOAD_FOLDER, TEMP_FOLDER]:
    if not os.path.exists(folder):
        try:
            os.makedirs(folder, mode=0o777, exist_ok=True)
            print(f"Created folder: {folder}")
        except Exception as e:
            print(f"Error creating folder {folder}: {str(e)}")

ALLOWED_EXTENSIONS = {'ifc', 'xlsx'}

def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def dms_to_decimal(dms):
    """Convert latitude/longitude in DMS format to decimal degrees."""
    degrees, minutes, seconds, direction = dms
    decimal = degrees + minutes / 60 + seconds / 3600
    if direction in ['S', 'W']:
        decimal = -decimal
    return decimal

def extract_georeferencing(ifc_file_path):
    """Extract georeferencing data from an IFC file."""
    try:
        ifc_file = ifcopenshell.open(ifc_file_path)
        site = ifc_file.by_type('IfcSite')[0]
        latitude = site.RefLatitude
        longitude = site.RefLongitude
        elevation = site.RefElevation

        latitude_decimal = dms_to_decimal(latitude)
        longitude_decimal = dms_to_decimal(longitude)
        return latitude_decimal, longitude_decimal, elevation
    except Exception as e:
        print(f"Error extracting georeferencing data: {e}")
        return None, None, None

def generate_map(latitude, longitude, elevation, output_file):
    """
    Génère une carte interactive avec Folium et sauvegarde le fichier HTML.

    Args:
        latitude (float): Latitude de l'emplacement.
        longitude (float): Longitude de l'emplacement.
        elevation (float): Altitude de l'emplacement.
        output_file (str): Chemin du fichier HTML à générer.

    Returns:
        bool: True si la carte a été générée avec succès, False sinon.
    """
    try:
        # Crée une carte centrée sur les coordonnées fournies
        m = folium.Map(location=[latitude, longitude], zoom_start=15)

        # Ajoute un marqueur à la position
        folium.Marker(
            [latitude, longitude],
            popup=f"Élévation : {elevation} m"
        ).add_to(m)

        # Sauvegarde la carte en tant que fichier HTML
        m.save(output_file)
        print(f"Carte générée et sauvegardée : {output_file}")
        return True
    except Exception as e:
        print(f"Erreur lors de la génération de la carte : {e}")
        return False

def load_element_types(file):
    element_types = {}
    df = pd.read_excel(file, sheet_name="Element_Types", engine='openpyxl')
    for index, row in df.iterrows():
        ifc_class = row["IFC_Class"]
        element_types[ifc_class] = ifc_class
    return element_types

def load_required_psets_and_params(file):
    required_psets_and_params = {}
    xls = pd.ExcelFile(file, engine='openpyxl')
    for sheet_name in xls.sheet_names:
        if sheet_name != "Element_Types":
            df = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
            for index, row in df.iterrows():
                ifc_class = row["IFC_Class"]
                param_name = row["Parametre"]
                param_type = row["Type"]
                if ifc_class not in required_psets_and_params:
                    required_psets_and_params[ifc_class] = {}
                if sheet_name not in required_psets_and_params[ifc_class]:
                    required_psets_and_params[ifc_class][sheet_name] = {}
                required_psets_and_params[ifc_class][sheet_name][param_name] = param_type
    return required_psets_and_params

def str_to_type(param_type_str):
    param_type_str = param_type_str.lower()
    if param_type_str.lower() == 'string':
        return str
    elif param_type_str.lower() == 'int':
        return int
    elif param_type_str.lower() == 'float':
        return float
    elif param_type_str == 'number':
        return float
    elif param_type_str.lower() == 'bool':
        return bool
    else:
        raise ValueError(f"Unsupported parameter type '{param_type_str}'")

def gray_empty_cells(worksheet):
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None or cell.value == "":
                cell.fill = gray_fill

def get_building_storey(ifc_file, element):
    try:
        if element is not None:
            if hasattr(element, 'ContainedInStructure') and element.ContainedInStructure:
                for rel in element.ContainedInStructure:
                    if rel.RelatingStructure.is_a("IfcBuildingStorey"):
                        return rel.RelatingStructure.Name if rel.RelatingStructure.Name else "Sans étage"
                    elif rel.RelatingStructure.is_a("IfcBuilding") or rel.RelatingStructure.is_a("IfcSite"):
                        return "Sans étage"
                    else:
                        return get_building_storey(ifc_file, rel.RelatingStructure)
        return "Sans étage"
    except Exception as e:
        print(f"Error getting building storey: {str(e)}")
        return "Sans étage"

def sort_floor_name(floor_name):
    if floor_name == "Sans étage":
        return "ZZZ"
    elif floor_name == "Sous-sol":
        return "AAA_-1"
    elif floor_name.startswith("R+"):
        try:
            return f"CCC_{int(floor_name[2:]):02d}"
        except ValueError:
            return f"DDD_{floor_name}"
    elif floor_name == "Rez-de-chaussée":
        return "BBB_00"
    elif floor_name.startswith("-"):
        try:
            return f"AAA_{int(floor_name.split('.')[0]):02d}"
        except ValueError:
            return f"DDD_{floor_name}"
    else:
        return f"DDD_{floor_name}"

def create_summary_sheet(workbook, total_elements, valid_elements, missing_elements, missing_psets, missing_params, floor_stats, elements_by_class, required_psets_and_params):
    # Créer l'onglet de résumé
    summary = workbook.create_sheet("Résumé", 0)
    
    # Styles
    title_font = Font(size=14, bold=True, color="FFFFFF")
    header_font = Font(size=12, bold=True)
    normal_font = Font(size=11)
    
    title_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ko_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    percent_format = '0.0%'
    
    def apply_style_to_range(start_cell, end_cell, font=None, fill=None, alignment=None, border=None, value=None, number_format=None):
        start_col = ord(start_cell[0].upper()) - ord('A') + 1
        end_col = ord(end_cell[0].upper()) - ord('A') + 1
        start_row = int(start_cell[1:])
        end_row = int(end_cell[1:])
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = summary.cell(row=row, column=col)
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
                if border:
                    cell.border = border
                if number_format:
                    cell.number_format = number_format
                if value and col == start_col and row == start_row:
                    cell.value = value
    
    # Titre principal
    apply_style_to_range('A1', 'E1', font=title_font, fill=title_fill, alignment=center_alignment, border=border, value="Rapport de validation IFC")
    summary.merge_cells('A1:E1')
    
    # Statistiques globales
    apply_style_to_range('A3', 'A3', font=header_font, fill=header_fill, value="Statistiques globales")
    
    stats = [
        ("Total d'éléments", total_elements),
        ("Éléments valides", valid_elements, valid_elements/total_elements if total_elements > 0 else 0),
        ("Éléments invalides", missing_elements, missing_elements/total_elements if total_elements > 0 else 0),
        ("PSet manquants", missing_psets),
        ("Paramètres manquants", missing_params)
    ]
    
    for i, stat in enumerate(stats, 4):
        label = stat[0]
        value = stat[1]
        percent = stat[2] if len(stat) > 2 else None
        
        apply_style_to_range(f'A{i}', f'A{i}', font=normal_font, alignment=left_alignment, border=border, value=label)
        apply_style_to_range(f'B{i}', f'B{i}', font=normal_font, alignment=center_alignment, border=border, value=value)
        if percent is not None:
            apply_style_to_range(f'C{i}', f'C{i}', font=normal_font, alignment=center_alignment, border=border, value=percent, number_format=percent_format)
    
    # Statistiques par étage
    apply_style_to_range('A7', 'A7', font=header_font, fill=header_fill, value="Répartition par étage")
    
    headers = ["Étage", "Total", "Valides", "Invalides", "Taux de validité"]
    for col, header in enumerate(headers, 1):
        apply_style_to_range(f'{chr(ord("A")+col-1)}8', f'{chr(ord("A")+col-1)}8', 
                           font=header_font, fill=header_fill, alignment=center_alignment, border=border, value=header)
    
    row = 9
    sorted_floors = sorted(floor_stats.items(), key=lambda x: sort_floor_name(x[0]))
    for floor, stats in sorted_floors:
        total = stats["valid"] + stats["invalid"]
        valid_rate = stats["valid"] / total if total > 0 else 0
        
        cells = [
            (floor, left_alignment),
            (total, center_alignment),
            (stats["valid"], center_alignment),
            (stats["invalid"], center_alignment),
            (valid_rate, center_alignment)
        ]
        
        for col, (value, alignment) in enumerate(cells, 1):
            cell_ref = f'{chr(ord("A")+col-1)}{row}'
            fill = None
            number_format = None
            
            if col == 5:  # Dernière colonne (taux)
                number_format = percent_format
                if valid_rate >= 0.8:
                    fill = ok_fill
                elif valid_rate < 0.5:
                    fill = ko_fill
            
            apply_style_to_range(cell_ref, cell_ref, 
                               font=normal_font, 
                               alignment=alignment, 
                               border=border, 
                               fill=fill, 
                               value=value,
                               number_format=number_format)
        
        row += 1
    
    # Statistiques par type
    type_title_row = row + 2
    apply_style_to_range(f'A{type_title_row}', f'A{type_title_row}', font=header_font, fill=header_fill, value="Répartition par type")
    
    headers = ["Type", "Total", "Valides", "Invalides", "Taux de validité"]
    header_row = type_title_row + 1
    for col, header in enumerate(headers, 1):
        apply_style_to_range(f'{chr(ord("A")+col-1)}{header_row}', f'{chr(ord("A")+col-1)}{header_row}',
                           font=header_font, fill=header_fill, alignment=center_alignment, border=border, value=header)
    
    row = header_row + 1
    sorted_types = sorted(elements_by_class.items())
    for element_type, stats in sorted_types:
        total = stats["total"]
        valid_rate = stats["valid"] / total if total > 0 else 0
        
        cells = [
            (element_type, left_alignment),
            (total, center_alignment),
            (stats["valid"], center_alignment),
            (stats["invalid"], center_alignment),
            (valid_rate, center_alignment)
        ]
        
        for col, (value, alignment) in enumerate(cells, 1):
            cell_ref = f'{chr(ord("A")+col-1)}{row}'
            fill = None
            number_format = None
            
            if col == 5:  # Dernière colonne (taux)
                number_format = percent_format
                if valid_rate >= 0.8:
                    fill = ok_fill
                elif valid_rate < 0.5:
                    fill = ko_fill
            
            apply_style_to_range(cell_ref, cell_ref,
                               font=normal_font,
                               alignment=alignment,
                               border=border,
                               fill=fill,
                               value=value,
                               number_format=number_format)
        
        row += 1
    
    # Ajuster la largeur des colonnes
    for col in range(1, 6):
        letter = get_column_letter(col)
        max_length = 0
        for cell in summary[letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        summary.column_dimensions[letter].width = adjusted_width
    
    # Figer les volets
    summary.freeze_panes = 'A2'
    
    # Activer l'onglet de résumé
    workbook.active = workbook.worksheets.index(summary)

# Mapping des éléments IFC vers les matériaux
IFC_TO_MATERIAL_MAPPING = {
    'IfcWall': 'Béton',
    'IfcWallStandardCase': 'Béton',
    'IfcSlab': 'Béton',
    'IfcFooting': 'Béton',
    'IfcBeam': 'Béton',
    'IfcColumn': 'Béton',
    'IfcStair': 'Béton',
    'IfcRailing': 'Acier',
    'IfcWindow': 'Verre',
    'IfcDoor': 'Bois',
    'IfcRoof': 'Béton',
    'IfcCovering': 'Isolation'
}

# Facteurs d'émission de CO2 par matériau
MATERIAL_CARBON_FACTORS = {
    'Béton': {
        'factor': 320,  # kg CO2e/m³
        'description': 'Béton armé standard',
        'source': 'Base INIES',
        'details': 'Inclut la production du ciment, des agrégats, le transport et la mise en œuvre'
    },
    'Acier': {
        'factor': 12000,  # kg CO2e/m³
        'description': 'Acier de construction',
        'source': 'Base INIES',
        'details': 'Inclut la production primaire, le laminage et le transport'
    },
    'Verre': {
        'factor': 2500,  # kg CO2e/m³
        'description': 'Double vitrage standard',
        'source': 'Base INIES',
        'details': 'Inclut la production du verre, l\'assemblage et le transport'
    },
    'Bois': {
        'factor': -750,  # kg CO2e/m³ (négatif car stockage de carbone)
        'description': 'Bois de construction',
        'source': 'Base INIES',
        'details': 'Bois local, inclut l\'exploitation forestière, le transport et la transformation'
    },
    'Isolation': {
        'factor': 100,  # kg CO2e/m³
        'description': 'Isolation thermique standard',
        'source': 'Base INIES',
        'details': 'Laine minérale, inclut la production, le transport et la mise en œuvre'
    }
}

def calculate_carbon_footprint(ifc_file, element):
    """Calcule l'empreinte carbone d'un élément IFC."""
    try:
        # Obtenir le type de matériau
        material = IFC_TO_MATERIAL_MAPPING.get(element.is_a(), 'Non spécifié')
        if material == 'Non spécifié':
            return 0
        
        # Obtenir le facteur d'émission
        material_data = MATERIAL_CARBON_FACTORS.get(material)
        if not material_data:
            return 0
        
        emission_factor = material_data['factor']
        
        # Calculer le volume ou la surface
        quantity = 0
        
        # Essayer d'abord d'obtenir les quantités via les propriétés
        if hasattr(element, 'IsDefinedBy'):
            for rel in element.IsDefinedBy:
                if rel.is_a('IfcRelDefinesByProperties'):
                    if rel.RelatingPropertyDefinition.is_a('IfcElementQuantity'):
                        for q in rel.RelatingPropertyDefinition.Quantities:
                            if q.is_a('IfcQuantityVolume'):
                                quantity = float(q.VolumeValue)
                                print(f"Volume found in properties for {element.is_a()}: {quantity:.2f} m³")
                                break
                            elif q.is_a('IfcQuantityArea'):
                                quantity = float(q.AreaValue)
                                print(f"Area found in properties for {element.is_a()}: {quantity:.2f} m²")
                                break
        
        # Si aucune quantité n'est trouvée, essayer de calculer à partir de la géométrie
        if quantity == 0:
            if hasattr(element, 'Representation'):
                try:
                    settings = ifcopenshell.geom.settings()
                    settings.set(settings.USE_PYTHON_OPENCASCADE, True)
                    shape = ifcopenshell.geom.create_shape(settings, element)
                    
                    if material in ['Verre', 'Isolation']:  # Pour les matériaux en surface
                        quantity = shape.surface_area
                        print(f"Surface area calculated for {element.is_a()}: {quantity:.2f} m²")
                        emission_factor = emission_factor / 10  # Convertir le facteur pour la surface
                    else:  # Pour les matériaux volumiques
                        quantity = shape.volume
                        print(f"Volume calculated for {element.is_a()}: {quantity:.2f} m³")
                except:
                    # Si le calcul géométrique échoue, estimer à partir des dimensions
                    if hasattr(element, 'OverallHeight') and hasattr(element, 'OverallWidth'):
                        if material in ['Verre', 'Isolation']:
                            quantity = float(element.OverallHeight) * float(element.OverallWidth)
                        else:
                            depth = 0.3  # Profondeur par défaut en mètres
                            if hasattr(element, 'OverallDepth'):
                                depth = float(element.OverallDepth)
                            quantity = float(element.OverallHeight) * float(element.OverallWidth) * depth
                        print(f"Quantity estimated from dimensions for {element.is_a()}: {quantity:.2f}")
                    else:
                        # Valeur par défaut si aucune autre méthode ne fonctionne
                        quantity = 1.0
                        print(f"Using default quantity for {element.is_a()}: {quantity:.2f}")
        
        # Calculer l'empreinte carbone
        carbon_footprint = quantity * emission_factor
        print(f"Carbon footprint for {element.is_a()}: {carbon_footprint:.2f} kg CO2e (Quantity: {quantity:.2f}, Factor: {emission_factor})")
        
        return carbon_footprint
    except Exception as e:
        print(f"Error calculating carbon footprint for {element.is_a()}: {str(e)}")
        return 0

def create_carbon_footprint_sheet(workbook, carbon_data):
    """Crée un onglet pour l'empreinte carbone dans le rapport Excel."""
    carbon_sheet = workbook.create_sheet("Empreinte_Carbone")
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    subheader_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Section 1: Bilan carbone du projet
    carbon_sheet.merge_cells('A1:F1')
    cell = carbon_sheet.cell(row=1, column=1)
    cell.value = "BILAN CARBONE DU PROJET"
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = header_fill
    
    # En-têtes du bilan
    headers = [
        "Type d'élément",
        "Matériau",
        "Volume/Surface (m³ ou m²)",
        "Facteur d'émission (kg CO2e/m³ ou kg CO2e/m²)",
        "Empreinte carbone (kg CO2e)",
        "% du total"
    ]
    
    # Ajuster la hauteur de la première ligne
    carbon_sheet.row_dimensions[1].height = 30
    
    for col, header in enumerate(headers, 1):
        cell = carbon_sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Ajuster la largeur des colonnes
        carbon_sheet.column_dimensions[get_column_letter(col)].width = 25
    
    # Ajuster la hauteur de la ligne d'en-tête
    carbon_sheet.row_dimensions[2].height = 45
    
    # Données du bilan
    row = 3
    row_start = row
    total_carbon = carbon_data['total']
    
    # Trier les éléments par empreinte carbone décroissante
    sorted_elements = sorted(
        carbon_data['by_type'].items(),
        key=lambda x: x[1],
        reverse=True
    )
    
    for element_type, carbon_value in sorted_elements:
        material = carbon_data['material_mapping'].get(element_type, "Non spécifié")
        material_data = carbon_data['material_factors'].get(material, {'factor': 0})
        
        # Calculer le volume/surface total pour ce type d'élément
        quantity = carbon_value / material_data['factor'] if material_data['factor'] != 0 else 0
        percentage = (carbon_value/total_carbon*100) if total_carbon != 0 else 0
        
        cells = [
            (element_type, "left"),
            (material, "left"),
            (f"{quantity:.2f}", "right"),
            (f"{material_data['factor']}", "right"),
            (f"{carbon_value:.2f}", "right"),
            (f"{percentage:.1f}%", "right")
        ]
        
        for col, (value, align) in enumerate(cells, 1):
            cell = carbon_sheet.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.font = normal_font
            cell.alignment = Alignment(horizontal=align, vertical="center")
        row += 1
    
    # Total
    row_end = row
    carbon_sheet.merge_cells(f'A{row}:C{row}')
    cell = carbon_sheet.cell(row=row, column=1)
    cell.value = "TOTAL"
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.fill = total_fill
    
    # Cellules du total
    for col in range(1, 7):
        cell = carbon_sheet.cell(row=row, column=col)
        cell.border = thin_border
        cell.fill = total_fill
        if col == 5:
            cell.value = f"{total_carbon:.2f}"
        elif col == 6:
            cell.value = "100%"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Section 2: Détails des matériaux
    row += 3
    carbon_sheet.merge_cells(f'A{row}:F{row}')
    cell = carbon_sheet.cell(row=row, column=1)
    cell.value = "DÉTAILS DES MATÉRIAUX ET FACTEURS D'ÉMISSION"
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = header_fill
    row += 1
    
    # En-têtes des détails
    material_headers = [
        "Matériau",
        "Description",
        "Facteur d'émission",
        "Source",
        "Détails",
        "Impact total (kg CO2e)"
    ]
    
    for col, header in enumerate(material_headers, 1):
        cell = carbon_sheet.cell(row=row, column=col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    
    material_row_start = row
    
    # Données des matériaux
    material_totals = defaultdict(float)
    for element_type, carbon_value in carbon_data['by_type'].items():
        material = carbon_data['material_mapping'].get(element_type, "Non spécifié")
        material_totals[material] += carbon_value
    
    # Trier les matériaux par impact total
    sorted_materials = sorted(
        [(mat, data, material_totals.get(mat, 0)) 
         for mat, data in carbon_data['material_factors'].items()],
        key=lambda x: x[2],
        reverse=True
    )
    
    for material, material_data, total in sorted_materials:
        cells = [
            (material, "left"),
            (material_data['description'], "left"),
            (f"{material_data['factor']} kg CO2e/m³", "right"),
            (material_data['source'], "left"),
            (material_data['details'], "left"),
            (f"{total:.2f}", "right")
        ]
        
        for col, (value, align) in enumerate(cells, 1):
            cell = carbon_sheet.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        row += 1
    
    material_row_end = row
    
    # Ajuster la hauteur des lignes pour le texte wrappé
    for row_idx in range(material_row_start, material_row_end):
        max_lines = max(
            len(str(carbon_sheet.cell(row=row_idx, column=col).value).split('\n'))
            for col in range(1, len(material_headers) + 1)
        )
        carbon_sheet.row_dimensions[row_idx].height = max(15 * max_lines, 30)
    
    # Ajouter les graphiques
    # 1. Graphique à barres pour l'empreinte carbone par type d'élément
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Empreinte carbone par élément"
    chart1.y_axis.title = "Empreinte carbone (kgCO2e)"
    chart1.x_axis.title = "Éléments"
    
    data = Reference(carbon_sheet, min_col=3, min_row=row_start-1, max_row=row_end-1)
    cats = Reference(carbon_sheet, min_col=1, min_row=row_start, max_row=row_end-1)
    
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    
    chart1.height = 15
    chart1.width = 25
    
    carbon_sheet.add_chart(chart1, "H2")
    
    # 2. Graphique circulaire pour la répartition par matériau
    chart2 = PieChart()
    chart2.title = "Répartition de l'empreinte carbone par matériau"
    chart2.style = 10
    
    pie_data = Reference(carbon_sheet, min_col=6, min_row=material_row_start-1, max_row=material_row_end-1)
    pie_labels = Reference(carbon_sheet, min_col=1, min_row=material_row_start, max_row=material_row_end-1)
    
    chart2.add_data(pie_data, titles_from_data=True)
    chart2.set_categories(pie_labels)
    
    # Configuration des étiquettes de données
    from openpyxl.chart.label import DataLabelList
    chart2.dataLabels = DataLabelList()
    chart2.dataLabels.showPercent = True
    chart2.dataLabels.showVal = False
    chart2.dataLabels.showCatName = True
    
    chart2.height = 15
    chart2.width = 25
    
    carbon_sheet.add_chart(chart2, "H20")

def process_files(temp_dir: str, ifc_file_path: str, excel_file_path: str, output_file_path: str):
    start_time = time.time()
    print(f"Starting analysis...")
    
    # Chargement des données
    element_types = load_element_types(excel_file_path)
    required_psets_and_params = load_required_psets_and_params(excel_file_path)
    ifc_file = ifcopenshell.open(ifc_file_path)
    model_name = os.path.basename(ifc_file_path)
    
    # Préfiltrage des éléments
    print(f"Loading elements... ({time.time() - start_time:.2f}s)")
    elements_by_class = {}
    missing_psets = 0
    missing_params = 0
    
    # Statistiques par étage
    floor_stats = defaultdict(lambda: {"valid": 0, "invalid": 0})
    
    # Analyse des éléments
    print(f"Analyzing elements... ({time.time() - start_time:.2f}s)")
    
    # Préparation du rapport Excel
    workbook = Workbook()
    details = workbook.active
    details.title = "Détails"
    
    # Styles pour l'onglet de détails
    header_font = Font(size=12, bold=True)
    normal_font = Font(size=11)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ko_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    no_fill = PatternFill(fill_type=None)
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes de l'onglet de détails
    headers = ["Type", "Étage", "ID", "Nom", "PSet", "Paramètre", "Valeur", "Statut"]
    for col, header in enumerate(headers, 1):
        cell = details.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
        cell.font = header_font
    
    row = 2
    total_carbon_footprint = 0
    carbon_footprint_by_type = defaultdict(float)
    carbon_footprint_by_floor = defaultdict(float)
    
    print("Analyzing elements and calculating carbon footprint...")
    for element in ifc_file.by_type('IfcElement'):
        if element.is_a() in element_types:
            if element.is_a() not in elements_by_class:
                elements_by_class[element.is_a()] = {"total": 0, "valid": 0, "invalid": 0}
            
            elements_by_class[element.is_a()]["total"] += 1
            floor = get_building_storey(ifc_file, element)
            element_psets = ifcopenshell.util.element.get_psets(element)
            element_valid = True
            has_missing_pset = False
            has_missing_param = False
            
            # Vérification des PSet et paramètres requis
            for pset_name, params in required_psets_and_params[element.is_a()].items():
                if pset_name not in element_psets:
                    element_valid = False
                    has_missing_pset = True
                    
                    cells = [
                        (1, element.is_a(), left_alignment),
                        (2, str(floor), left_alignment),
                        (3, element.GlobalId, center_alignment),
                        (4, getattr(element, 'Name', ''), left_alignment),
                        (5, pset_name, left_alignment),
                        (6, "TOUS", center_alignment),
                        (7, "MANQUANT", center_alignment),
                        (8, "KO", center_alignment)
                    ]
                    
                    for col, value, alignment in cells:
                        cell = details.cell(row=row, column=col)
                        cell.value = value
                        cell.alignment = alignment
                        cell.border = border
                        cell.font = normal_font
                        cell.fill = ko_fill if col == 8 else no_fill
                    
                    row += 1
                else:
                    for param_name, param_type in params.items():
                        actual_value = element_psets[pset_name].get(param_name)
                        
                        if actual_value is None:
                            element_valid = False
                            has_missing_param = True
                            
                            cells = [
                                (1, element.is_a(), left_alignment),
                                (2, str(floor), left_alignment),
                                (3, element.GlobalId, center_alignment),
                                (4, getattr(element, 'Name', ''), left_alignment),
                                (5, pset_name, left_alignment),
                                (6, param_name, left_alignment),
                                (7, "MANQUANT", center_alignment),
                                (8, "KO", center_alignment)
                            ]
                            
                            for col, value, alignment in cells:
                                cell = details.cell(row=row, column=col)
                                cell.value = value
                                cell.alignment = alignment
                                cell.border = border
                                cell.font = normal_font
                                cell.fill = ko_fill if col == 8 else no_fill
                            
                            row += 1
                        else:
                            try:
                                param_type_class = str_to_type(param_type)
                                param_type_class(actual_value)
                                
                                cells = [
                                    (1, element.is_a(), left_alignment),
                                    (2, str(floor), left_alignment),
                                    (3, element.GlobalId, center_alignment),
                                    (4, getattr(element, 'Name', ''), left_alignment),
                                    (5, pset_name, left_alignment),
                                    (6, param_name, left_alignment),
                                    (7, str(actual_value), left_alignment),
                                    (8, "OK", center_alignment)
                                ]
                                
                                for col, value, alignment in cells:
                                    cell = details.cell(row=row, column=col)
                                    cell.value = value
                                    cell.alignment = alignment
                                    cell.border = border
                                    cell.font = normal_font
                                    cell.fill = ok_fill if col == 8 else no_fill
                                
                                row += 1
                            except (ValueError, TypeError):
                                element_valid = False
                                has_missing_param = True
                                
                                cells = [
                                    (1, element.is_a(), left_alignment),
                                    (2, str(floor), left_alignment),
                                    (3, element.GlobalId, center_alignment),
                                    (4, getattr(element, 'Name', ''), left_alignment),
                                    (5, pset_name, left_alignment),
                                    (6, param_name, left_alignment),
                                    (7, f"{actual_value} (attendu: {param_type})", left_alignment),
                                    (8, "KO", center_alignment)
                                ]
                                
                                for col, value, alignment in cells:
                                    cell = details.cell(row=row, column=col)
                                    cell.value = value
                                    cell.alignment = alignment
                                    cell.border = border
                                    cell.font = normal_font
                                    cell.fill = ko_fill if col == 8 else no_fill
                                
                                row += 1
            
            # Mise à jour des statistiques
            if element_valid:
                elements_by_class[element.is_a()]["valid"] += 1
                floor_stats[floor]["valid"] += 1
            else:
                elements_by_class[element.is_a()]["invalid"] += 1
                floor_stats[floor]["invalid"] += 1
                if has_missing_pset:
                    missing_psets += 1
                if has_missing_param:
                    missing_params += 1
            
            # Calcul de l'empreinte carbone
            try:
                carbon_footprint = calculate_carbon_footprint(ifc_file, element)
                print(f"Carbon footprint for {element.is_a()}: {carbon_footprint:.2f} kg CO2e")
                total_carbon_footprint += carbon_footprint
                carbon_footprint_by_type[element.is_a()] += carbon_footprint
                carbon_footprint_by_floor[floor] += carbon_footprint
            except Exception as e:
                print(f"Error calculating carbon footprint for {element.is_a()}: {str(e)}")
    
    # Statistiques globales
    total_elements = sum(stats["total"] for stats in elements_by_class.values())
    valid_elements = sum(stats["valid"] for stats in elements_by_class.values())
    invalid_elements = sum(stats["invalid"] for stats in elements_by_class.values())
    
    # Créer l'onglet de résumé
    create_summary_sheet(workbook, total_elements, valid_elements, invalid_elements, missing_psets, missing_params, floor_stats, elements_by_class, required_psets_and_params)
    
    print(f"Creating carbon footprint sheet with total: {total_carbon_footprint:.2f} kg CO2e")
    print(f"Carbon footprint by type: {dict(carbon_footprint_by_type)}")
    print(f"Carbon footprint by floor: {dict(carbon_footprint_by_floor)}")
    
    # Ajouter la feuille d'empreinte carbone
    carbon_data = {
        'total': total_carbon_footprint,
        'by_type': dict(carbon_footprint_by_type),
        'by_floor': dict(carbon_footprint_by_floor),
        'material_mapping': IFC_TO_MATERIAL_MAPPING,
        'material_factors': MATERIAL_CARBON_FACTORS
    }
    create_carbon_footprint_sheet(workbook, carbon_data)
    
    # Ajuster les largeurs des colonnes
    for ws in workbook.worksheets:
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # Figer les volets dans l'onglet de détails
    details.freeze_panes = 'A2'
    
    # Sauvegarder le fichier
    print(f"Saving file... ({time.time() - start_time:.2f}s)")
    workbook.save(output_file_path)
    print(f"Analysis completed in {time.time() - start_time:.2f}s")
    
    # Préparation des données pour le dashboard
    sorted_floors = sorted(floor_stats.items(), key=lambda x: sort_floor_name(x[0]))
    return {
        "total_elements": total_elements,
        "valid_elements": valid_elements,
        "missing_elements": invalid_elements,
        "missing_psets": missing_psets,
        "missing_params": missing_params,
        "floors": [{"name": floor, "valid": stats["valid"], "invalid": stats["invalid"]} for floor, stats in sorted_floors],
        "carbon_footprint": {
            "total": total_carbon_footprint,
            "by_type": dict(carbon_footprint_by_type),
            "by_floor": dict(carbon_footprint_by_floor)
        }
    }

@app.route('/upload', methods=['POST'])
def upload():
    try:
        print("Starting upload...")

        # Vérification des fichiers dans la requête
        if 'ifc_file' not in request.files or 'excel_file' not in request.files:
            print("Missing files in request")
            return jsonify({"error": "Missing IFC or Excel file"}), 400
        
        ifc_file = request.files['ifc_file']
        excel_file = request.files['excel_file']

        # Vérification des noms de fichiers
        if ifc_file.filename == '' or excel_file.filename == '':
            print("Empty filenames")
            return jsonify({"error": "No selected file"}), 400
        
        # Vérification des types de fichiers
        if not allowed_file(ifc_file.filename) or not allowed_file(excel_file.filename):
            print("Invalid file types")
            return jsonify({"error": "Invalid file type"}), 400

        # Génération d'un ID unique pour l'analyse
        analysis_id = str(uuid.uuid4())
        analysis_dir = os.path.join(TEMP_FOLDER, analysis_id)
        os.makedirs(analysis_dir, exist_ok=True)

        # Sauvegarde des fichiers téléchargés
        ifc_path = os.path.join(analysis_dir, secure_filename(ifc_file.filename))
        excel_path = os.path.join(analysis_dir, secure_filename(excel_file.filename))
        output_path = os.path.join(analysis_dir, f'output_{os.path.splitext(ifc_file.filename)[0]}.xlsx')

        print(f"Saving files to: {ifc_path}, {excel_path}")
        ifc_file.save(ifc_path)
        excel_file.save(excel_path)

        # Étape 1 : Analyse des fichiers
        print("Starting analysis...")
        results = process_files(analysis_dir, ifc_path, excel_path, output_path)

        # Étape 2 : Extraction des données de géoréférencement
        print("Extracting georeferencing data...")
        latitude, longitude, elevation = extract_georeferencing(ifc_path)
        if latitude is None or longitude is None:
            print("Failed to extract georeferencing data")
            return jsonify({"error": "Failed to extract georeferencing data"}), 500

        # Étape 3 : Génération de la carte
        map_file = os.path.join(GEOMAP_FOLDER, f"{analysis_id}_map.html")
        if not generate_map(latitude, longitude, elevation, map_file):
            print("Failed to generate map")
            return jsonify({"error": "Failed to generate map"}), 500

        # Ajout des données de géoréférencement au résultat
        results.update({
            "analysis_id": analysis_id,
            "latitude": latitude,
            "longitude": longitude,
            "elevation": elevation,
            "map_url": f"/geomaps/{os.path.basename(map_file)}"
        })

        print("Upload and processing completed successfully")
        return jsonify(results)

    except Exception as e:
        print(f"Unexpected error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/download/<analysis_id>')
def download(analysis_id):
    try:
        analysis_dir = os.path.join(TEMP_FOLDER, analysis_id)
        if not os.path.exists(analysis_dir):
            return jsonify({"error": "Analysis not found"}), 404
        
        # Trouver le fichier Excel de sortie
        output_files = [f for f in os.listdir(analysis_dir) if f.startswith('output_') and f.endswith('.xlsx')]
        if not output_files:
            return jsonify({"error": "Output file not found"}), 404
        
        output_path = os.path.join(analysis_dir, output_files[0])
        return send_file(output_path, as_attachment=True, download_name=output_files[0])
        
    except Exception as e:
        print(f"Error during download: {str(e)}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    # Configuration du port via variable d'environnement pour Vercel
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)
